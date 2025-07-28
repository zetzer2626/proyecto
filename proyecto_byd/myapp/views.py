from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView,View
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from .models import Proyecto,Solicitud,Contactos,Tramitacion,Pago_tramitacion,Listado,Insumos,Evento,Enlace,Boleta,PagoCliente, DetallePago
from .forms import ProyectoForm, UserRegistrationForm,SolicitudForm,TramitacionForm,PagoTramitacionForm,ContactoForm,InsumoForm,PagoClienteForm, DetallePagoForm, BuscarPagoForm
from django.shortcuts import render, redirect,get_object_or_404
from django.contrib.auth import authenticate, login
from django.contrib.auth.views import LoginView, LogoutView
from django.contrib import messages
from django.http import JsonResponse
from django.db.models import Q
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import User
from django.contrib.auth.models import Permission
from django.db.models import Sum
from django.http import HttpResponse
from django.utils.dateparse import parse_date
from django.utils import timezone
from datetime import datetime
from django.template.loader import render_to_string
import calendar
from datetime import datetime
from django.core.paginator import Paginator
import openpyxl
from openpyxl.styles import Font, Border, Side
from django.db.models import Sum, Q
from django.contrib import messages
from decimal import Decimal
from django.views.decorators.http import require_POST



@method_decorator(login_required, name='dispatch')
class ProyectoListView(ListView):
    model = Proyecto
    template_name = 'myapp/proyecto_list.html'
    context_object_name = 'proyectos'
    ordering = ['-id']  # Orden descendente por ID

    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get('q')
        estado = self.request.GET.get('estado')
        proceso = self.request.GET.get('proceso')
        año = self.request.GET.get('año')

        # Búsqueda en varios campos
        if query:
            queryset = queryset.filter(
                Q(nombre__icontains=query) |
                Q(proyecto__icontains=query) |
                Q(direccion__icontains=query)
            )

        # Filtro por estado del proyecto
        if estado:
            queryset = queryset.filter(estado_proyecto=estado)

        # Filtro por proceso
        if proceso:
            queryset = queryset.filter(proceso=proceso)

        # Filtro por año
        if año:
            queryset = queryset.filter(año=año)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Contadores generales para las tarjetas
        context['total_proyectos'] = Proyecto.objects.count()
        context['total_aprobados'] = Proyecto.objects.filter(estado_proyecto='aprobado').count()
        context['total_no_concretados'] = Proyecto.objects.filter(estado_proyecto='no_concretado').count()
        
        # Contador para "Trabajar en Proyecto" ← NUEVO CONTADOR
        context['total_trabajando'] = Proyecto.objects.filter(estado_proyecto='trabajando').count()
        
        # Actualizar el contador de ingresados para incluir todos los tipos
        context['total_ingresado'] = Proyecto.objects.filter(
            estado_proyecto__in=['ingreso_dom', 'ingresado_sag', 'ingresado_minvu', 
                               'ingresado_monumento', 'observado', 'rechazado']
        ).count()
        
        context['total_observados'] = Proyecto.objects.filter(estado_proyecto='observado').count()
        
        return context


@method_decorator(login_required, name='dispatch')
class ProyectoCreateView(CreateView):
    model = Proyecto
    form_class = ProyectoForm
    template_name = 'myapp/proyecto_create.html'
    success_url = reverse_lazy('proyecto-list')

@method_decorator(login_required, name='dispatch')
class ProyectoUpdateView(UpdateView):
    model = Proyecto
    form_class = ProyectoForm
    template_name = 'myapp/proyecto_edit.html'
    success_url = reverse_lazy('proyecto-list')

@method_decorator(login_required, name='dispatch')
class ProyectoDetailView(DetailView):
    model = Proyecto
    template_name = 'myapp/proyecto_detail.html'
    context_object_name = 'proyecto'

@method_decorator(login_required, name='dispatch')
class ProyectoDeleteView(DeleteView):
    model = Proyecto
    template_name = 'myapp/proyecto_confirm_delete.html'
    success_url = reverse_lazy('proyecto-list')
    
    def delete(self, request, *args, **kwargs):
        print(f"Intentando eliminar proyecto con ID: {kwargs.get('pk')}")
        try:
            return super().delete(request, *args, **kwargs)
        except Exception as e:
            print(f"Error al eliminar proyecto: {e}")
            messages.error(request, f"Error al eliminar el proyecto: {e}")
            return redirect('proyecto-list')



# Vista para el login
def custom_login(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('proyecto-list')  # Redirige a la lista de proyectos
            else:
                messages.error(request, "Usuario o contraseña incorrectos.")
        else:
            messages.error(request, "Formulario inválido.")
    else:
        form = AuthenticationForm()
    return render(request, 'registration/login.html', {'form': form})

# Vista para el logout
def custom_logout(request):
    logout(request)
    return redirect('login')  # Redirige al login

# Vista para el registro de usuarios
def register(request):
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data['password'])
            user.save()
            messages.success(request, "Cuenta creada con éxito.")
            return redirect('login')  # Redirige al login después de registrarse
    else:
        form = UserRegistrationForm()
    return render(request, 'registration/register.html', {'form': form})



def analisis(request):
    from django.db.models import Count, Q
    from datetime import datetime, timedelta
    import json
    
    # Estadísticas básicas
    total_proyectos = Proyecto.objects.count()
    proyectos_completados = Proyecto.objects.filter(estado_proyecto='aprobado').count()
    proyectos_no_concretados = Proyecto.objects.filter(estado_proyecto='no_concretado').count()
    proyectos_trabajando = Proyecto.objects.filter(estado_proyecto='trabajando').count()
    proyectos_observados = Proyecto.objects.filter(estado_proyecto='observado').count()
    proyectos_congelados = Proyecto.objects.filter(estado_proyecto='congelado').count()
    
    # Estadísticas por proceso
    procesos_terminados = Proyecto.objects.filter(proceso='terminado').count()
    procesos_pendientes = Proyecto.objects.filter(proceso='pendiente').count()
    procesos_no_aplica = Proyecto.objects.filter(proceso='no aplica').count()
    
    # Estadísticas por año
    proyectos_por_año = Proyecto.objects.values('año').annotate(
        count=Count('id')
    ).order_by('año')
    
    # Proyectos del último mes
    fecha_limite = datetime.now() - timedelta(days=30)
    proyectos_recientes = Proyecto.objects.filter(
        Q(estado_proyecto='aprobado') | Q(estado_proyecto='trabajando')
    ).count()
    
    # Distribución por estado
    estados_data = {
        'Aprobados': proyectos_completados,
        'No Concretados': proyectos_no_concretados,
        'Trabajando': proyectos_trabajando,
        'Observados': proyectos_observados,
        'Congelados': proyectos_congelados
    }
    
    # Distribución por proceso
    procesos_data = {
        'Terminados': procesos_terminados,
        'Pendientes': procesos_pendientes,
        'No Aplica': procesos_no_aplica
    }
    
    # Datos para gráficos
    chart_data = {
        'estados': {
            'labels': list(estados_data.keys()),
            'data': list(estados_data.values()),
            'colors': ['#28a745', '#dc3545', '#ffc107', '#17a2b8', '#6c757d']
        },
        'procesos': {
            'labels': list(procesos_data.keys()),
            'data': list(procesos_data.values()),
            'colors': ['#28a745', '#ffc107', '#6c757d']
        },
        'por_año': {
            'labels': [str(item['año']) for item in proyectos_por_año],
            'data': [item['count'] for item in proyectos_por_año],
            'colors': ['#007bff']
        }
    }
    
    # Calcular porcentajes
    porcentaje_completados = (proyectos_completados / total_proyectos * 100) if total_proyectos > 0 else 0
    porcentaje_no_concretados = (proyectos_no_concretados / total_proyectos * 100) if total_proyectos > 0 else 0
    porcentaje_trabajando = (proyectos_trabajando / total_proyectos * 100) if total_proyectos > 0 else 0
    
    # Proyectos más recientes (últimos 5)
    proyectos_recientes_list = Proyecto.objects.order_by('-id')[:5]
    
    # Top 5 años con más proyectos
    top_años = Proyecto.objects.values('año').annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    context = {
        'total_proyectos': total_proyectos,
        'proyectos_completados': proyectos_completados,
        'proyectos_no_concretados': proyectos_no_concretados,
        'proyectos_trabajando': proyectos_trabajando,
        'proyectos_observados': proyectos_observados,
        'proyectos_congelados': proyectos_congelados,
        'procesos_terminados': procesos_terminados,
        'procesos_pendientes': procesos_pendientes,
        'procesos_no_aplica': procesos_no_aplica,
        'proyectos_recientes': proyectos_recientes,
        'porcentaje_completados': round(porcentaje_completados, 1),
        'porcentaje_no_concretados': round(porcentaje_no_concretados, 1),
        'porcentaje_trabajando': round(porcentaje_trabajando, 1),
        'proyectos_recientes_list': proyectos_recientes_list,
        'top_años': top_años,
        'chart_data': json.dumps(chart_data),
        'estados_data': estados_data,
        'procesos_data': procesos_data,
    }
    
    return render(request, 'myapp/analisis.html', context)


###############AVISTAS DEL MODELO SOLICITRUD######################

@method_decorator(login_required, name='dispatch')
class SolicitudListView(ListView):
    model = Solicitud
    template_name = 'solicitud/solicitud_list.html'
    context_object_name = 'solicitudes'
    ordering = ['-id']  # Orden descendente id

    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get('q')
        estado = self.request.GET.get('estado')

        # Filtro de búsqueda por solicitante, nombre y dirección
        if query:
            queryset = queryset.filter(
                Q(solicitante__icontains=query) | 
                Q(nombre__icontains=query) |
                Q(direccion__icontains=query) |
                Q(nombre_documento__icontains=query)
            )
        
        # Filtro por estado de solicitud
        if estado:
            if estado == "completado":
                queryset = queryset.filter(fecha_solicitud__isnull=False, fecha_recepcion__isnull=False)
            elif estado == "en_proceso":
                queryset = queryset.filter(fecha_solicitud__isnull=False, fecha_recepcion__isnull=True)
            elif estado == "pendiente":
                queryset = queryset.filter(fecha_solicitud__isnull=True)

        return queryset

    def get_context_data(self, **kwargs):
        try:
            context = super().get_context_data(**kwargs)
            
            # Obtener los contadores por estado
            context['completados_count'] = Solicitud.objects.filter(fecha_solicitud__isnull=False, fecha_recepcion__isnull=False).count()
            context['en_proceso_count'] = Solicitud.objects.filter(fecha_solicitud__isnull=False, fecha_recepcion__isnull=True).count()
            context['pendientes_count'] = Solicitud.objects.filter(fecha_solicitud__isnull=True).count()

            # Mantener los parámetros de búsqueda en el contexto para que se mantengan cuando se navegue entre páginas
            context['query'] = self.request.GET.get('q', '')
            context['estado'] = self.request.GET.get('estado', '')

            return context
        except Exception as e:
            # En caso de error, devolver contexto básico
            context = super().get_context_data(**kwargs)
            context['completados_count'] = 0
            context['en_proceso_count'] = 0
            context['pendientes_count'] = 0
            context['query'] = self.request.GET.get('q', '')
            context['estado'] = self.request.GET.get('estado', '')
            return context
    
@method_decorator(login_required, name='dispatch')
class SolicitudCreateView(CreateView):
    model = Solicitud
    form_class = SolicitudForm
    template_name = 'solicitud/create_solicitud.html'
    success_url = reverse_lazy('solicitud-list')
    
    def form_valid(self, form):
        try:
            # Save the form
            self.object = form.save()
            
            # Check if it's an AJAX request
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Solicitud creada exitosamente',
                    'redirect_url': self.success_url
                })
            
            # For non-AJAX requests, redirect
            return redirect(self.success_url)
        except Exception as e:
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': f'Error al crear la solicitud: {str(e)}',
                    'errors': {}
                })
            raise
    
    def form_invalid(self, form):
        # Check if it's an AJAX request
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': 'Error al crear la solicitud',
                'errors': form.errors
            })
        
        return super().form_invalid(form)

@method_decorator(login_required, name='dispatch')
class SolicitudUpdateView(UpdateView):
    model = Solicitud
    form_class = SolicitudForm
    template_name = 'solicitud/edit_solicitud.html'
    success_url = reverse_lazy('solicitud-list')
    
    def form_valid(self, form):
        try:
            # Save the form
            self.object = form.save()
            
            # Check if it's an AJAX request
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Solicitud actualizada exitosamente',
                    'redirect_url': self.success_url
                })
            
            # For non-AJAX requests, redirect
            return redirect(self.success_url)
        except Exception as e:
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': f'Error al actualizar la solicitud: {str(e)}',
                    'errors': {}
                })
            raise
    
    def form_invalid(self, form):
        # Check if it's an AJAX request
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': 'Error al actualizar la solicitud',
                'errors': form.errors
            })
        
        return super().form_invalid(form)

@method_decorator(login_required, name='dispatch')
class SolicitudDetailView(DetailView):
    model = Solicitud
    template_name = 'solicitud/detalle_solicitud.html'
    context_object_name = 'solicitud'

@method_decorator(login_required, name='dispatch')
class SolicitudDeleteView(DeleteView):
    model = Solicitud
    template_name = 'solicitud/delete_solicitud.html'
    success_url = reverse_lazy('solicitud-list')
    
    def delete(self, request, *args, **kwargs):
        response = super().delete(request, *args, **kwargs)
        
        # Check if it's an AJAX request
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': 'Solicitud eliminada exitosamente',
                'redirect_url': self.success_url
            })
        
        return response



@method_decorator(login_required, name='dispatch') # Vista de lista (index)  
class ContactosListView(ListView):  
    model = Contactos  
    form_class = ContactoForm
    template_name = 'contacto/contactos_list.html'  
    context_object_name = 'contactos'  
    ordering = ['-id']  # Orden descendente id
    
    def get_queryset(self):  
        queryset = super().get_queryset()  
        query = self.request.GET.get('q')  
        filtro_tipo = self.request.GET.get('tipo')  

        # Filtrar por búsqueda de nombre, correo o teléfono  
        if query:  
            queryset = queryset.filter(  
                Q(nombre__icontains=query) |   
                Q(correo__icontains=query) |   
                Q(telefono__icontains=query)  
            )  
        
        # Filtrar por tipo de contacto  
        if filtro_tipo:  
            queryset = queryset.filter(tipo=filtro_tipo)  # Cambiado a 'tipo'  

        return queryset  

    def get_context_data(self, **kwargs):  
        context = super().get_context_data(**kwargs)  
        
        # Calcular los contadores  
        total_contactos = Contactos.objects.count()  
        total_clientes = Contactos.objects.filter(tipo='CLIENTE').count()  # Cambiado a 'tipo'  
        total_profesionales = Contactos.objects.filter(tipo='PROFESIONAL').count()  # Cambiado a 'tipo'  
        total_otros = Contactos.objects.filter(tipo='OTROS').count()  # Cambiado a 'tipo'  

        # Debug: imprimir información
        print(f"DEBUG: Total contactos: {total_contactos}")
        print(f"DEBUG: Total clientes: {total_clientes}")
        print(f"DEBUG: Total profesionales: {total_profesionales}")
        print(f"DEBUG: Total otros: {total_otros}")

        # Agregar los contadores al contexto  
        context['total_contactos'] = total_contactos  
        context['total_clientes'] = total_clientes  
        context['total_profesionales'] = total_profesionales  
        context['total_otros'] = total_otros  
        
        return context  

# Vista de detalle (ver)
@method_decorator(login_required, name='dispatch')
class ContactosDetailView(DetailView):
    model = Contactos
    template_name = 'contacto/contactos_detail.html'
    context_object_name = 'contacto'

# Vista de creación (crear)
@method_decorator(login_required, name='dispatch')
class ContactosCreateView(CreateView):
    model = Contactos
    template_name = 'contacto/contactos_create.html'
    fields = '__all__'
    success_url = reverse_lazy('contacto-list')  # Redirigir a la lista tras la creación
    
    def form_valid(self, form):
        try:
            # Save the form
            self.object = form.save()
            
            # Check if it's an AJAX request
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Contacto creado exitosamente',
                    'redirect_url': self.success_url
                })
            
            # For non-AJAX requests, redirect
            return redirect(self.success_url)
        except Exception as e:
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': f'Error al crear el contacto: {str(e)}',
                    'errors': {}
                })
            raise
    
    def form_invalid(self, form):
        # Check if it's an AJAX request
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': 'Error al crear el contacto',
                'errors': form.errors
            })
        
        return super().form_invalid(form)

# Vista de actualización (editar)
@method_decorator(login_required, name='dispatch')
class ContactosUpdateView(UpdateView):
    model = Contactos
    template_name = 'contacto/contactos_edit.html'
    fields = '__all__'
    success_url = reverse_lazy('contacto-list')
    
    def form_valid(self, form):
        try:
            # Save the form
            self.object = form.save()
            
            # Check if it's an AJAX request
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Contacto actualizado exitosamente',
                    'redirect_url': self.success_url
                })
            
            # For non-AJAX requests, redirect
            return redirect(self.success_url)
        except Exception as e:
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': f'Error al actualizar el contacto: {str(e)}',
                    'errors': {}
                })
            raise
    
    def form_invalid(self, form):
        # Check if it's an AJAX request
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': 'Error al actualizar el contacto',
                'errors': form.errors
            })
        
        return super().form_invalid(form)

# Vista de eliminación (eliminar)
@method_decorator(login_required, name='dispatch')
class ContactosDeleteView(DeleteView):
    model = Contactos
    template_name = 'contacto/contactos_delete.html'
    success_url = reverse_lazy('contacto-list')
    
    def delete(self, request, *args, **kwargs):
        response = super().delete(request, *args, **kwargs)
        
        # Check if it's an AJAX request
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': 'Contacto eliminado exitosamente',
                'redirect_url': self.success_url
            })
        
        return response


################## Vista de TRAMITACIONES ##########################33

@method_decorator(login_required, name='dispatch')
class TramitacionListView(ListView):
    model = Tramitacion
    template_name = 'tramitacion/tramitacion_list.html'
    context_object_name = 'tramitaciones'
    ordering = ['-id']  # Orden descendente id

    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.GET.get('q')
        estado = self.request.GET.get('estado')
        año = self.request.GET.get('año')
        fecha_desde = self.request.GET.get('fecha_desde')

        # Filtro de búsqueda general
        if query:
            queryset = queryset.filter(
                Q(presupuesto__icontains=query) | 
                Q(nombre__icontains=query) |
                Q(direccion__icontains=query) |
                Q(nota__icontains=query)
            )
        
        # Filtro por año
        if año:
            queryset = queryset.filter(año=año)
        
        # Filtro por fecha desde
        if fecha_desde:
            queryset = queryset.filter(fecha_recepcion__gte=fecha_desde)
        
        # Filtro por estado basado en 'fecha_recepcion'
        if estado:
            if estado == "completado":
                queryset = queryset.filter(fecha_recepcion__isnull=False)
            elif estado == "pendiente":
                queryset = queryset.filter(fecha_recepcion__isnull=True)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Obtener los contadores por estado basado en 'fecha_recepcion'
        context['completados_count'] = Tramitacion.objects.filter(fecha_recepcion__isnull=False).count()
        context['sin_fecha_count'] = Tramitacion.objects.filter(fecha_recepcion__isnull=True).count()

        # Mantener los parámetros de búsqueda en el contexto
        context['query'] = self.request.GET.get('q', '')
        context['estado'] = self.request.GET.get('estado', '')
        context['año'] = self.request.GET.get('año', '')
        context['fecha_desde'] = self.request.GET.get('fecha_desde', '')

        return context

    
@method_decorator(login_required, name='dispatch')
@method_decorator(login_required, name='dispatch')
class TramitacionCreateView(CreateView):
    model = Tramitacion
    form_class = TramitacionForm
    template_name = 'tramitacion/create_tramitacion.html'
    success_url = reverse_lazy('tramitacion-list')

    def form_valid(self, form):
        try:
            # Guardar la tramitación
            self.object = form.save()
            
            # Verificar si es una solicitud AJAX
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Tramitación creada exitosamente',
                    'id': self.object.id
                })
            
            # Para solicitudes normales, usar el comportamiento por defecto
            return super().form_valid(form)
        except Exception as e:
            print(f"Error en form_valid (create): {e}")
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': f'Error al crear la tramitación: {str(e)}'
                }, status=500)
            else:
                raise

    def form_invalid(self, form):
        try:
            # Verificar si es una solicitud AJAX
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': 'Error al crear la tramitación. Verifique los datos.',
                    'errors': form.errors
                }, status=400)
            
            # Para solicitudes normales, usar el comportamiento por defecto
            return super().form_invalid(form)
        except Exception as e:
            print(f"Error en form_invalid (create): {e}")
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': f'Error al procesar el formulario: {str(e)}'
                }, status=500)
            else:
                raise

@method_decorator(login_required, name='dispatch')
class TramitacionUpdateView(UpdateView):
    model = Tramitacion
    form_class = TramitacionForm
    template_name = 'tramitacion/edit_tramitacion.html'
    success_url = reverse_lazy('tramitacion-list')

    def form_valid(self, form):
        try:
            # Guardar la tramitación
            self.object = form.save()
            
            # Verificar si es una solicitud AJAX
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Tramitación actualizada exitosamente',
                    'id': self.object.id
                })
            
            # Para solicitudes normales, usar el comportamiento por defecto
            return super().form_valid(form)
        except Exception as e:
            print(f"Error en form_valid: {e}")
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': f'Error al actualizar la tramitación: {str(e)}'
                }, status=500)
            else:
                raise

    def form_invalid(self, form):
        try:
            # Verificar si es una solicitud AJAX
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': 'Error al actualizar la tramitación. Verifique los datos.',
                    'errors': form.errors
                }, status=400)
            
            # Para solicitudes normales, usar el comportamiento por defecto
            return super().form_invalid(form)
        except Exception as e:
            print(f"Error en form_invalid: {e}")
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': f'Error al procesar el formulario: {str(e)}'
                }, status=500)
            else:
                raise

@method_decorator(login_required, name='dispatch')
class TramitacionDetailView(DetailView):
    model = Tramitacion
    template_name = 'tramitacion/detalle_tramitacion.html'
    context_object_name = 'tramitacion'

@method_decorator(login_required, name='dispatch')
class TramitacionDeleteView(DeleteView):
    model = Tramitacion
    template_name = 'tramitacion/delete_tramitacion.html'
    success_url = reverse_lazy('tramitacion-list')

    def delete(self, request, *args, **kwargs):
        try:
            self.object = self.get_object()
            
            # Verificar si es una solicitud AJAX
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                try:
                    self.object.delete()
                    return JsonResponse({
                        'success': True,
                        'message': 'Tramitación eliminada exitosamente'
                    })
                except Exception as e:
                    print(f"Error al eliminar tramitación: {e}")
                    return JsonResponse({
                        'success': False,
                        'message': f'Error al eliminar la tramitación: {str(e)}'
                    }, status=400)
            
            # Para solicitudes normales, usar el comportamiento por defecto
            return super().delete(request, *args, **kwargs)
            
        except Exception as e:
            print(f"Error general en delete: {e}")
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': f'Error al procesar la solicitud: {str(e)}'
                }, status=500)
            else:
                # Para solicitudes normales, re-raise la excepción
                raise



#################################--FIN--######################################################


####################################---INGRESO Y EGRESO--##########################################################
@method_decorator(login_required, name='dispatch')
class PagoTramitacionListView(ListView):
    model = Pago_tramitacion
    template_name = 'pago_tramitacion/pago_tramitacion_list.html'
    context_object_name = 'pagos'

    def get_queryset(self):
        queryset = super().get_queryset()

        # Obtener los filtros desde los parámetros GET
        fecha_inicio = self.request.GET.get('fecha_inicio')
        fecha_fin = self.request.GET.get('fecha_fin')
        tipo_pago = self.request.GET.get('tipo_pago')

        # Aplicar filtros al queryset
        if fecha_inicio:
            queryset = queryset.filter(fecha__gte=fecha_inicio)
        if fecha_fin:
            queryset = queryset.filter(fecha__lte=fecha_fin)
        if tipo_pago:
            queryset = queryset.filter(tipo_pago=tipo_pago)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        queryset = self.get_queryset()

        # Calcular totales
        total_ingresos = queryset.aggregate(Sum('ingreso'))['ingreso__sum'] or 0
        total_egresos = queryset.aggregate(Sum('egreso'))['egreso__sum'] or 0
        saldo = total_ingresos - total_egresos

        context.update({
            'total_ingresos_filtrados': total_ingresos,
            'total_egresos_filtrados': total_egresos,
            'saldo_filtrado': saldo,
        })
        return context
    
# Crear registro
@method_decorator(login_required, name='dispatch')
class PagoTramitacionCreateView(CreateView):
    model = Pago_tramitacion
    template_name = 'pago_tramitacion/pago_tramitacion_create.html'
    form_class = PagoTramitacionForm
    success_url = reverse_lazy('pago_tramitacion_list')


# Editar registro
@method_decorator(login_required, name='dispatch')
class PagoTramitacionUpdateView(UpdateView):
    model = Pago_tramitacion
    template_name = 'pago_tramitacion/pago_tramitacion_edit.html'
    form_class = PagoTramitacionForm
    success_url = reverse_lazy('pago_tramitacion_list')

# Detalle de un registro
@method_decorator(login_required, name='dispatch')
class PagoTramitacionDetailView(DetailView):
    model = Pago_tramitacion
    template_name = 'pago_tramitacion/pago_tramitacion_detail.html'
    context_object_name = 'pago'

# Eliminar registro
@method_decorator(login_required, name='dispatch')
class PagoTramitacionDeleteView(DeleteView):
    model = Pago_tramitacion
    template_name = 'pago_tramitacion/pago_tramitacion_confirm_delete.html'
    success_url = reverse_lazy('pago_tramitacion_list')
    
#################################--FIN--######################################################


#################################listado######################################################

# Vista para listar todos los objetos
@method_decorator(login_required, name='dispatch')
class ListadoListView(ListView):
    model = Listado
    template_name = 'listado/listado_list.html'
    context_object_name = 'listados'
    ordering = ['-id']  # Orden descendente id
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Obtener todos los listados para estadísticas
        listados = Listado.objects.all()
        
        # Estadísticas
        context['listados_con_pagos'] = listados.filter(listado__isnull=False).exclude(listado='').count()
        context['listados_con_drive'] = listados.filter(drive__isnull=False).exclude(drive='').count()
        
        # Año actual y listados del año actual
        from datetime import datetime
        año_actual = datetime.now().year
        context['año_actual'] = año_actual
        context['listados_año_actual'] = listados.filter(año=str(año_actual)).count()
        
        # Años disponibles para filtros
        años_disponibles = listados.values_list('año', flat=True).distinct().order_by('-año')
        context['años_disponibles'] = años_disponibles
        
        return context

# Vista para crear un nuevo objeto
@method_decorator(login_required, name='dispatch')
class ListadoCreateView(CreateView):
    model = Listado
    template_name = 'listado/listado_create.html'
    fields = '__all__'
    success_url = reverse_lazy('listado_list')  # Redirige a la lista después de crear

# Vista para editar un objeto existente
@method_decorator(login_required, name='dispatch')
class ListadoUpdateView(UpdateView):
    model = Listado
    template_name = 'listado/listado_edit.html'
    fields = '__all__'
    success_url = reverse_lazy('listado_list')

# Vista para ver los detalles de un objeto
@method_decorator(login_required, name='dispatch')
class ListadoDetailView(DetailView):
    model = Listado
    template_name = 'listado/listado_detail.html'
    context_object_name = 'listado'

# Eliminar registro
@method_decorator(login_required, name='dispatch')
class ListadoDeleteView(DeleteView):
    model = Listado
    template_name = 'listado/listado_delete.html'
    success_url = reverse_lazy('listado_list')

# Vistas AJAX para modales
@login_required
def crear_listado_ajax(request):
    if request.method == 'POST':
        try:
            # Crear el listado
            listado = Listado.objects.create(
                presupuesto=request.POST.get('presupuesto'),
                año=request.POST.get('año'),
                nombre=request.POST.get('nombre'),
                gestion=request.POST.get('gestion'),
                nota=request.POST.get('nota'),
                listado=request.POST.get('listado') or None,
                drive=request.POST.get('drive') or None
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Listado "{listado.presupuesto}" creado exitosamente',
                'listado_id': listado.id
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al crear listado: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Método no permitido'})

@login_required
def editar_listado_ajax(request, pk):
    try:
        listado = Listado.objects.get(pk=pk)
    except Listado.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Listado no encontrado'})
    
    if request.method == 'POST':
        try:
            # Actualizar el listado
            listado.presupuesto = request.POST.get('presupuesto')
            listado.año = request.POST.get('año')
            listado.nombre = request.POST.get('nombre')
            listado.gestion = request.POST.get('gestion')
            listado.nota = request.POST.get('nota')
            listado.listado = request.POST.get('listado') or None
            listado.drive = request.POST.get('drive') or None
            listado.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Listado "{listado.presupuesto}" actualizado exitosamente'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al actualizar listado: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Método no permitido'})

@login_required
def eliminar_listado_ajax(request, pk):
    try:
        listado = Listado.objects.get(pk=pk)
    except Listado.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Listado no encontrado'})
    
    if request.method == 'POST':
        try:
            presupuesto = listado.presupuesto
            listado.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'Listado "{presupuesto}" eliminado exitosamente'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al eliminar listado: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Método no permitido'})
#################################--FIN--######################################################


##########################################--INSUMOS--#####################################################################33

@method_decorator(login_required, name='dispatch')
class InsumosListView(ListView):  
    model = Insumos  
    template_name = 'insumo/insumo_list.html'  # Cambia esto según el nombre de tu template  
    context_object_name = 'insumos'  # Nombre del contexto a usar en el template  
    ordering = ['-id']  # Orden descendente id

@method_decorator(login_required, name='dispatch')
class InsumosDetailView(DetailView):  
    model = Insumos  
    template_name = 'insumo/insumo_detail.html'  # Cambia esto según el nombre de tu template  
    context_object_name = 'insumo'  

@method_decorator(login_required, name='dispatch')
class InsumosCreateView(CreateView):  
    model = Insumos  
    template_name = 'insumo/insumo_create.html'  # Cambia esto según el nombre de tu template  
    form_class = InsumoForm
    success_url = reverse_lazy('insumo_list')  # Asegúrate de tener esta URL definida en tu urls.py  

@method_decorator(login_required, name='dispatch')
class InsumosUpdateView(UpdateView):  
    model = Insumos  
    template_name = 'insumo/insumo_edit.html'  # Usa el mismo que para Crear  
    form_class = InsumoForm
    success_url = reverse_lazy('insumo_list')  # Asegúrate de tener esta URL definida en tu urls.py  

@method_decorator(login_required, name='dispatch')
class InsumosDeleteView(DeleteView):  
    model = Insumos  
    template_name = 'insumo/insumo_delete.html'  # Cambia esto según el nombre de tu template  
    success_url = reverse_lazy('insumo_list')  # Asegúrate de tener esta URL definida en tu urls.py  
    #################################--FIN--######################################################



###############Calendario#####################3

def calendario(request):
    # Obtener el año y el mes desde la URL o usar los valores actuales
    anio = int(request.GET.get('anio', datetime.now().year))
    mes = int(request.GET.get('mes', datetime.now().month))

    # Calcular mes anterior y siguiente
    mes_anterior = mes - 1 if mes > 1 else 12
    anio_anterior = anio if mes > 1 else anio - 1
    mes_siguiente = mes + 1 if mes < 12 else 1
    anio_siguiente = anio if mes < 12 else anio + 1

    # Generar el calendario
    cal = calendar.Calendar()
    dias_del_mes = cal.itermonthdays2(anio, mes)  # Devuelve (día, día de la semana)
    dias = []
    semana = []

    for dia, dia_semana in dias_del_mes:
        if dia == 0:  # Día fuera del mes actual
            semana.append({"dia": "", "eventos": []})
        else:
            fecha_actual = datetime(anio, mes, dia).date()
            eventos = Evento.objects.filter(fecha=fecha_actual)
            semana.append({"dia": dia, "eventos": eventos})
        if len(semana) == 7:  # Semana completa
            dias.append(semana)
            semana = []

    if semana:  # Agregar la última semana incompleta
        dias.append(semana)

    # Manejo del formulario para agregar eventos
    if request.method == "POST" and "agregar_evento" in request.POST:
        titulo = request.POST.get("titulo")
        descripcion = request.POST.get("descripcion")
        fecha_str = request.POST.get("fecha")

        try:
            # Convertir la fecha recibida al tipo Date
            fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()

            # Crear el evento
            Evento.objects.create(titulo=titulo, descripcion=descripcion, fecha=fecha)
            return redirect(f"/cuentas/calendario/?anio={anio}&mes={mes}")
        except ValueError:
            # En caso de error al convertir la fecha
            return redirect(f"/cuentas/calendario/?anio={anio}&mes={mes}&error=Fecha inválida")

    # Manejo de edición de eventos
    if request.method == "POST" and "editar_evento" in request.POST:
        evento_id = request.POST.get("evento_id")
        titulo = request.POST.get("titulo")
        descripcion = request.POST.get("descripcion")

        evento = get_object_or_404(Evento, id=evento_id)
        evento.titulo = titulo
        evento.descripcion = descripcion
        evento.save()

        return redirect(f"/cuentas/calendario/?anio={anio}&mes={mes}")

    # Manejo de eliminación de eventos
    if request.method == "POST" and "eliminar_evento" in request.POST:
        evento_id = request.POST.get("evento_id")

        evento = get_object_or_404(Evento, id=evento_id)
        evento.delete()

        return redirect(f"/cuentas/calendario/?anio={anio}&mes={mes}")

    # Manejo de marcar/desmarcar el ticket (completado)
    if request.method == "POST" and "marcar_ticket" in request.POST:
        evento_id = request.POST.get("evento_id")
        evento = get_object_or_404(Evento, id=evento_id)
        evento.completado = True  # Marcar como completado
        evento.save()
        return redirect(f"/cuentas/calendario/?anio={anio}&mes={mes}")

    if request.method == "POST" and "desmarcar_ticket" in request.POST:
        evento_id = request.POST.get("evento_id")
        evento = get_object_or_404(Evento, id=evento_id)
        evento.completado = False  # Desmarcar como completado
        evento.save()
        return redirect(f"/cuentas/calendario/?anio={anio}&mes={mes}")

    # Contexto para la plantilla
    contexto = {
        'anio': anio,
        'mes': mes,
        'nombre_mes': calendar.month_name[mes],
        'dias': dias,
        'mes_anterior': mes_anterior,
        'anio_anterior': anio_anterior,
        'mes_siguiente': mes_siguiente,
        'anio_siguiente': anio_siguiente,
    }
    return render(request, 'actividades/calendario.html', contexto)
#################################--FIN--######################################################



#################################--VISTA DE ENLACE--######################################################
# Vista para listar los enlaces
class EnlaceListView(ListView):
    model = Enlace
    template_name = 'enlace/enlace_list.html'
    context_object_name = 'enlaces'
    ordering = ['-id']  # Orden descendente id

    def get_queryset(self):
        query = self.request.GET.get('q', '')  # Captura el parámetro de búsqueda
        object_list = Enlace.objects.filter(
            Q(titulo__icontains=query) | Q(description__icontains=query)  # Filtra por título o descripción
        ).order_by('-id')  # Ordena por el más reciente
        return object_list

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        query = self.request.GET.get('q', '')  # Incluye el término de búsqueda en el contexto
        context['query'] = query
        context['total_enlaces'] = self.get_queryset().count()  # Total de enlaces
        context['current_page_count'] = len(context['enlaces'])  # Enlaces en la página actual
        
        # Estadísticas adicionales
        context['enlaces_con_pdf'] = Enlace.objects.filter(documento_pdf__isnull=False).exclude(documento_pdf='').count()
        context['enlaces_con_link'] = Enlace.objects.filter(link__isnull=False).exclude(link='').count()
        context['enlaces_con_avatar'] = Enlace.objects.filter(avatar__isnull=False).exclude(avatar='').count()
        
        return context

# Vista para crear un nuevo enlace
class EnlaceCreateView(CreateView):
    model = Enlace
    template_name = 'enlace/enlace_create.html'  # Cambia según la ubicación de tu plantilla
    fields = '__all__'
    success_url = reverse_lazy('enlace_list')

# Vista para actualizar un enlace existente
class EnlaceUpdateView(UpdateView):
    model = Enlace
    template_name = 'enlace/enlace_edit.html'  # Cambia según la ubicación de tu plantilla
    fields = '__all__'
    success_url = reverse_lazy('enlace_list')

# Vista para eliminar un enlace
class EnlaceDeleteView(DeleteView):
    model = Enlace
    template_name = 'enlace/enlace_delete.html'  # Cambia según la ubicación de tu plantilla
    success_url = reverse_lazy('enlace_list')

# Vista para ver detalles de un enlace
class EnlaceDetailView(DetailView):
    model = Enlace
    template_name = 'enlace/enlace_detail.html'
    context_object_name = 'enlace'

# Vistas AJAX para enlaces
@login_required
def crear_enlace_ajax(request):
    if request.method == 'POST':
        try:
            enlace = Enlace.objects.create(
                titulo=request.POST.get('titulo'),
                description=request.POST.get('description'),
                link=request.POST.get('link') or None,
                documento_pdf=request.FILES.get('documento_pdf'),
                avatar=request.FILES.get('avatar')
            )
            return JsonResponse({
                'success': True,
                'message': f'Enlace "{enlace.titulo}" creado exitosamente',
                'enlace_id': enlace.id
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al crear enlace: {str(e)}'
            })
    return JsonResponse({'success': False, 'message': 'Método no permitido'})

@login_required
def editar_enlace_ajax(request, pk):
    if request.method == 'POST':
        try:
            enlace = Enlace.objects.get(pk=pk)
            enlace.titulo = request.POST.get('titulo')
            enlace.description = request.POST.get('description')
            enlace.link = request.POST.get('link') or None
            
            # Manejar archivos
            if 'documento_pdf' in request.FILES:
                enlace.documento_pdf = request.FILES['documento_pdf']
            if 'avatar' in request.FILES:
                enlace.avatar = request.FILES['avatar']
            
            enlace.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Enlace "{enlace.titulo}" actualizado exitosamente'
            })
        except Enlace.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Enlace no encontrado'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al editar enlace: {str(e)}'
            })
    return JsonResponse({'success': False, 'message': 'Método no permitido'})

@login_required
def eliminar_enlace_ajax(request, pk):
    if request.method == 'POST':
        try:
            enlace = Enlace.objects.get(pk=pk)
            titulo = enlace.titulo
            enlace.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'Enlace "{titulo}" eliminado exitosamente'
            })
        except Enlace.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Enlace no encontrado'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al eliminar enlace: {str(e)}'
            })
    return JsonResponse({'success': False, 'message': 'Método no permitido'})
#################################--FIN--######################################################


class BoletaListView(ListView):
    model = Boleta
    template_name = 'boleta/boleta_list.html'
    context_object_name = 'boletas'
    ordering = ['-id']  # Orden descendente por id
    
class BoletaCreateView(CreateView):
    model = Boleta
    template_name = 'boleta/boleta_create.html'
    fields = '__all__'
    success_url = reverse_lazy('boleta_list')

class BoletaUpdateView(UpdateView):
    model = Boleta
    template_name = 'boleta/boleta_edit.html'
    fields = '__all__'
    success_url = reverse_lazy('boleta_list')

class BoletaDeleteView(DeleteView):
    model = Boleta
    template_name = 'boleta/boleta_delete.html'
    success_url = reverse_lazy('boleta_list')


@method_decorator(login_required, name='dispatch')
class BoletaDetailView(DetailView):
    model = Boleta
    template_name = 'boleta/boleta_detail.html'
    context_object_name = 'boleta'


############ IMPORTACIONES DE EXCEL DE PROYECTOS #######################
def exportar_excel(request):
    # Crear un libro de Excel y una hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Proyectos'

    # Encabezados de la tabla
    headers = ['Pres', 'Año', 'Nombre', 'Proyecto', 'Dirección', 'Estado Proyecto', 'Detalle']
    ws.append(headers)

    # Estilo para los encabezados (negrita y borde)
    font = Font(bold=True)
    border = Border(
        top=Side(border_style="thin"),
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )
    
    # Aplicar estilo de encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = font
        cell.border = border

    # Obtener los datos de tu modelo
    proyectos = Proyecto.objects.all()

    # Agregar los datos de la tabla
    for proyecto in proyectos:
        ws.append([
            proyecto.presupuesto if proyecto.presupuesto else 'No disponible',
            proyecto.año if proyecto.año else 'No disponible',
            proyecto.nombre if proyecto.nombre else 'No disponible',
            proyecto.proyecto if proyecto.proyecto else 'No disponible',
            proyecto.direccion if proyecto.direccion else 'No disponible',
            proyecto.estado_proyecto if proyecto.estado_proyecto else 'No disponible',
            proyecto.detalle if proyecto.detalle else 'No disponible',
        ])

    # Ajustar el ancho de las columnas según el contenido
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Obtener la letra de la columna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=proyectos_exportados.xlsx'
    
    # Guardar el archivo en la respuesta
    wb.save(response)
    return response

################################### FIN  ##########################################


def exportar_tramitacion_excel(request):
    # Crear un libro de Excel y una hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Tramitaciones'

    # Encabezados de la tabla
    headers = [
        'Presupuesto', 'Año', 'Nombre', 'Dirección', 'Nota', 'N° Subdiv. Aprobada', 'Estado'
    ]
    ws.append(headers)

    # Estilo para los encabezados (negrita y borde)
    font = Font(bold=True)
    border = Border(
        top=Side(border_style="thin"),
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )
    
    # Aplicar estilo de encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = font
        cell.border = border

    # Obtener los datos de tu modelo
    tramitaciones = Tramitacion.objects.all()

    # Agregar los datos de la tabla
    for tramitacion in tramitaciones:
        ws.append([
            tramitacion.presupuesto,
            tramitacion.año,
            tramitacion.nombre,
            tramitacion.direccion,
            tramitacion.nota if tramitacion.nota else 'No disponible',
            'N° Subdiv. Aprobada' if tramitacion.documento_tramitacion else 'No disponible',  # Asume que 'N° Subdiv. Aprobada' es algo específico
            'Estado' if tramitacion.fecha_solicitud else 'Pendiente'  # Asume un estado simple basado en si hay o no fecha_solicitud
        ])

    # Ajustar el ancho de las columnas según el contenido
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Obtener la letra de la columna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=tramitaciones_exportadas.xlsx'
    
    # Guardar el archivo en la respuesta
    wb.save(response)
    return response




# ============ VISTAS PARA PAGOS DE CLIENTES ============

@method_decorator(login_required, name='dispatch')
class PagoClienteListView(ListView):
    model = PagoCliente
    template_name = 'pagos_cliente/pago_cliente_list.html'
    context_object_name = 'pagos'
    ordering = ['-fecha_creacion']
    paginate_by = 20

    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Obtener parámetros de filtro
        cliente = self.request.GET.get('cliente')
        estado = self.request.GET.get('estado')
        proyecto_id = self.request.GET.get('proyecto')
        fecha_desde = self.request.GET.get('fecha_desde')
        fecha_hasta = self.request.GET.get('fecha_hasta')

        # Aplicar filtros
        if cliente:
            queryset = queryset.filter(
                Q(cliente_nombre__icontains=cliente) |
                Q(proyecto__nombre__icontains=cliente)
            )
        
        if estado:
            queryset = queryset.filter(estado_pago=estado)
        
        if proyecto_id:
            queryset = queryset.filter(proyecto_id=proyecto_id)
        
        if fecha_desde:
            queryset = queryset.filter(fecha_creacion__date__gte=fecha_desde)
        
        if fecha_hasta:
            queryset = queryset.filter(fecha_creacion__date__lte=fecha_hasta)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Estadísticas generales
        queryset = self.get_queryset()
        context['total_pagos'] = queryset.count()
        context['total_monto'] = queryset.aggregate(Sum('monto_total'))['monto_total__sum'] or 0
        context['total_pagado'] = queryset.aggregate(Sum('monto_pagado'))['monto_pagado__sum'] or 0
        context['total_pendiente'] = context['total_monto'] - context['total_pagado']
        
        # Calcular promedio por cliente
        if context['total_pagos'] > 0:
            context['promedio_por_cliente'] = context['total_monto'] / context['total_pagos']
        else:
            context['promedio_por_cliente'] = 0
        
        # Contadores por estado
        context['pendientes_count'] = queryset.filter(estado_pago='pendiente').count()
        context['parciales_count'] = queryset.filter(estado_pago='parcial').count()
        context['completos_count'] = queryset.filter(estado_pago='completo').count()
        context['vencidos_count'] = queryset.filter(estado_pago='vencido').count()
        
        # Formulario de búsqueda
        context['form_buscar'] = BuscarPagoForm(self.request.GET)
        
        return context


@method_decorator(login_required, name='dispatch')
class PagoClienteCreateView(CreateView):
    model = PagoCliente
    form_class = PagoClienteForm
    template_name = 'pagos_cliente/pago_cliente_create.html'
    success_url = reverse_lazy('pago_cliente_list')

    def form_valid(self, form):
        messages.success(self.request, 'Pago de cliente creado exitosamente.')
        return super().form_valid(form)


@method_decorator(login_required, name='dispatch')
class PagoClienteUpdateView(UpdateView):
    model = PagoCliente
    form_class = PagoClienteForm
    template_name = 'pagos_cliente/pago_cliente_edit.html'
    success_url = reverse_lazy('pago_cliente_list')

    def form_valid(self, form):
        messages.success(self.request, 'Pago de cliente actualizado exitosamente.')
        return super().form_valid(form)


@method_decorator(login_required, name='dispatch')
class PagoClienteDetailView(DetailView):
    model = PagoCliente
    template_name = 'pagos_cliente/pago_cliente_detail.html'
    context_object_name = 'pago'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['detalles_pago'] = self.object.detalles_pago.all().order_by('-fecha_pago')
        context['form_detalle_pago'] = DetallePagoForm()
        return context


@method_decorator(login_required, name='dispatch')
class PagoClienteDeleteView(DeleteView):
    model = PagoCliente
    template_name = 'pagos_cliente/pago_cliente_delete.html'
    success_url = reverse_lazy('pago_cliente_list')
    context_object_name = 'object'

    def delete(self, request, *args, **kwargs):
        try:
            pago = self.get_object()
            nombre_cliente = pago.cliente_nombre
            
            # Eliminar el objeto
            response = super().delete(request, *args, **kwargs)
            
            # Mensaje de éxito
            messages.success(request, f'Pago de {nombre_cliente} eliminado exitosamente.')
            return response
            
        except Exception as e:
            messages.error(request, f'Error al eliminar el pago: {str(e)}')
            return redirect('pago_cliente_detail', pk=self.kwargs['pk'])


# ============ VISTAS PARA DETALLES DE PAGO ============

@login_required
def crear_detalle_pago(request, pago_id):
    pago_cliente = get_object_or_404(PagoCliente, id=pago_id)
    
    if request.method == 'POST':
        form = DetallePagoForm(request.POST, request.FILES, pago_cliente=pago_cliente)
        if form.is_valid():
            detalle = form.save(commit=False)
            detalle.pago_cliente = pago_cliente
            detalle.registrado_por = request.user
            
            # Verificar que no se exceda el monto pendiente
            if detalle.monto_pago > pago_cliente.monto_pendiente:
                messages.error(request, f'El monto no puede exceder el pendiente: ${pago_cliente.monto_pendiente:,.2f}')
                return redirect('pago_cliente_detail', pk=pago_id)
            
            detalle.save()
            messages.success(request, f'Pago de ${detalle.monto_pago:,.2f} registrado exitosamente.')
            return redirect('pago_cliente_detail', pk=pago_id)
        else:
            messages.error(request, 'Error en el formulario. Verifique los datos.')
    
    return redirect('pago_cliente_detail', pk=pago_id)


@login_required
def editar_detalle_pago(request, detalle_id):
    detalle = get_object_or_404(DetallePago, id=detalle_id)
    pago_cliente = detalle.pago_cliente
    
    if request.method == 'POST':
        form = DetallePagoForm(request.POST, request.FILES, instance=detalle, pago_cliente=pago_cliente)
        if form.is_valid():
            form.save()
            messages.success(request, 'Detalle de pago actualizado exitosamente.')
            return redirect('pago_cliente_detail', pk=pago_cliente.id)
    else:
        form = DetallePagoForm(instance=detalle, pago_cliente=pago_cliente)
    
    return render(request, 'pagos_cliente/detalle_pago_edit.html', {
        'form': form,
        'detalle': detalle,
        'pago_cliente': pago_cliente
    })


@login_required
def eliminar_detalle_pago(request, detalle_id):
    detalle = get_object_or_404(DetallePago, id=detalle_id)
    pago_cliente = detalle.pago_cliente
    
    if request.method == 'POST':
        detalle.delete()
        messages.success(request, 'Detalle de pago eliminado exitosamente.')
    
    return redirect('pago_cliente_detail', pk=pago_cliente.id)


# ============ REPORTES Y EXPORTACIONES ============

@login_required
def reporte_pagos_cliente(request):
    """Vista para generar reportes de pagos"""
    
    # Obtener filtros
    form_buscar = BuscarPagoForm(request.GET)
    pagos = PagoCliente.objects.all()
    
    if form_buscar.is_valid():
        cliente = form_buscar.cleaned_data.get('cliente')
        estado = form_buscar.cleaned_data.get('estado')
        proyecto = form_buscar.cleaned_data.get('proyecto')
        fecha_desde = form_buscar.cleaned_data.get('fecha_desde')
        fecha_hasta = form_buscar.cleaned_data.get('fecha_hasta')
        
        if cliente:
            pagos = pagos.filter(
                Q(cliente_nombre__icontains=cliente) |
                Q(proyecto__nombre__icontains=cliente)
            )
        if estado:
            pagos = pagos.filter(estado_pago=estado)
        if proyecto:
            pagos = pagos.filter(proyecto=proyecto)
        if fecha_desde:
            pagos = pagos.filter(fecha_creacion__date__gte=fecha_desde)
        if fecha_hasta:
            pagos = pagos.filter(fecha_creacion__date__lte=fecha_hasta)
    
    # Estadísticas del reporte
    estadisticas = {
        'total_clientes': pagos.values('cliente_nombre').distinct().count(),
        'total_monto': pagos.aggregate(Sum('monto_total'))['monto_total__sum'] or 0,
        'total_pagado': pagos.aggregate(Sum('monto_pagado'))['monto_pagado__sum'] or 0,
        'total_pendiente': 0,
        'por_estado': {}
    }
    
    estadisticas['total_pendiente'] = estadisticas['total_monto'] - estadisticas['total_pagado']
    
    # Estadísticas por estado
    for estado, nombre in PagoCliente.ESTADO_PAGO_CHOICES:
        count = pagos.filter(estado_pago=estado).count()
        monto = pagos.filter(estado_pago=estado).aggregate(Sum('monto_total'))['monto_total__sum'] or 0
        estadisticas['por_estado'][estado] = {'count': count, 'monto': monto, 'nombre': nombre}
    
    context = {
        'pagos': pagos.order_by('-fecha_creacion'),
        'form_buscar': form_buscar,
        'estadisticas': estadisticas,
    }
    
# ============ API PARA OBTENER DATOS DEL PROYECTO ============

@login_required
def obtener_datos_proyecto(request, proyecto_id):
    """API para obtener datos del proyecto para auto-completar formulario"""
    try:
        proyecto = get_object_or_404(Proyecto, id=proyecto_id)
        data = {
            'nombre': proyecto.nombre or '',
            'telefono': proyecto.telefono or '',
            'correo': proyecto.correo or '',
            'direccion': proyecto.direccion or '',
        }
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def actualizar_notas_proyecto(request, proyecto_id):
    """Vista para actualizar las notas internas de un proyecto via AJAX"""
    if request.method == 'POST':
        try:
            import json
            proyecto = get_object_or_404(Proyecto, id=proyecto_id)
            data = json.loads(request.body)
            notas_internas = data.get('notas_internas', '').strip()
            
            proyecto.notas_internas = notas_internas
            proyecto.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Notas actualizadas exitosamente',
                'notas_internas': notas_internas
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

@login_required
def buscar_proyectos_ajax(request):
    """Vista AJAX para buscar proyectos por nombre o correo"""
    query = request.GET.get('q', '').strip()
    
    if not query or len(query) < 2:
        return JsonResponse({
            'success': False,
            'message': 'Debe ingresar al menos 2 caracteres para buscar'
        })
    
    try:
        # Buscar proyectos que coincidan con nombre o correo
        proyectos = Proyecto.objects.filter(
            Q(nombre__icontains=query) | 
            Q(correo__icontains=query)
        )[:10]  # Limitar a 10 resultados
        
        resultados = []
        for proyecto in proyectos:
            resultados.append({
                'id': proyecto.id,
                'nombre': proyecto.nombre,
                'correo': proyecto.correo or '',
                'telefono': proyecto.telefono or '',
                'direccion': proyecto.direccion or '',
                'presupuesto': proyecto.presupuesto or '',
                'estado': proyecto.estado_proyecto or '',
                'texto_busqueda': f"{proyecto.nombre} - {proyecto.correo or 'Sin correo'}"
            })
        
        return JsonResponse({
            'success': True,
            'resultados': resultados,
            'total': len(resultados)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al buscar proyectos: {str(e)}'
        })

@login_required
def obtener_datos_proyecto_ajax(request, proyecto_id):
    """Vista AJAX para obtener datos completos de un proyecto"""
    try:
        proyecto = get_object_or_404(Proyecto, id=proyecto_id)
        
        return JsonResponse({
            'success': True,
            'proyecto': {
                'id': proyecto.id,
                'nombre': proyecto.nombre,
                'correo': proyecto.correo or '',
                'telefono': proyecto.telefono or '',
                'direccion': proyecto.direccion or '',
                'presupuesto': proyecto.presupuesto or '',
                'estado': proyecto.estado_proyecto or '',
                'descripcion': proyecto.descripcion or '',
                'año': proyecto.año or '',
            }
        })
        
    except Proyecto.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Proyecto no encontrado'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al obtener datos del proyecto: {str(e)}'
        })

@login_required
def crear_pago_cliente_ajax(request):
    """Vista AJAX para crear un nuevo PagoCliente"""
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'message': 'Método no permitido'
        })
    
    try:
        # Obtener datos del formulario
        proyecto_id = request.POST.get('proyecto_id')
        cliente_nombre = request.POST.get('cliente_nombre')
        cliente_telefono = request.POST.get('cliente_telefono', '')
        cliente_correo = request.POST.get('cliente_correo', '')
        monto_total = request.POST.get('monto_total')
        fecha_vencimiento = request.POST.get('fecha_vencimiento', '')
        descripcion = request.POST.get('descripcion', '')
        observaciones = request.POST.get('observaciones', '')
        
        # Validaciones básicas
        if not proyecto_id or not cliente_nombre or not monto_total:
            return JsonResponse({
                'success': False,
                'message': 'Faltan campos obligatorios'
            })
        
        # Obtener proyecto
        proyecto = get_object_or_404(Proyecto, id=proyecto_id)
        
        # Crear PagoCliente
        pago_cliente = PagoCliente.objects.create(
            proyecto=proyecto,
            cliente_nombre=cliente_nombre,
            cliente_telefono=cliente_telefono,
            cliente_correo=cliente_correo,
            monto_total=int(monto_total),
            fecha_vencimiento=fecha_vencimiento if fecha_vencimiento else None,
            descripcion=descripcion,
            observaciones=observaciones
        )
        
        # Actualizar estado automáticamente
        pago_cliente.actualizar_estado()
        
        return JsonResponse({
            'success': True,
            'message': 'Pago de cliente creado exitosamente',
            'pago_cliente': {
                'id': pago_cliente.id,
                'cliente_nombre': pago_cliente.cliente_nombre,
                'monto_total': pago_cliente.monto_total,
                'monto_pagado': pago_cliente.monto_pagado,
                'monto_pendiente': pago_cliente.monto_pendiente,
                'estado_pago': pago_cliente.estado_pago,
                'fecha_creacion': pago_cliente.fecha_creacion.strftime('%d/%m/%Y'),
                'porcentaje_pagado': pago_cliente.porcentaje_pagado
            }
        })
        
    except ValueError as e:
        return JsonResponse({
            'success': False,
            'message': f'Error en el formato de datos: {str(e)}'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al crear pago de cliente: {str(e)}'
        })

@login_required
def crear_detalle_pago_ajax(request):
    """Vista AJAX para crear un nuevo DetallePago"""
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'message': 'Método no permitido'
        })
    
    try:
        # Obtener datos del formulario
        pago_cliente_id = request.POST.get('pago_cliente_id')
        fecha_pago = request.POST.get('fecha_pago')
        monto_pago = request.POST.get('monto_pago')
        forma_pago = request.POST.get('forma_pago')
        numero_referencia = request.POST.get('numero_referencia', '')
        notas = request.POST.get('notas', '')
        
        print(f"DEBUG - Datos recibidos: pago_cliente_id={pago_cliente_id}, fecha_pago={fecha_pago}, monto_pago={monto_pago}, forma_pago={forma_pago}")
        
        # Validaciones básicas
        if not pago_cliente_id or not fecha_pago or not monto_pago or not forma_pago:
            print(f"DEBUG - Faltan campos: pago_cliente_id={bool(pago_cliente_id)}, fecha_pago={bool(fecha_pago)}, monto_pago={bool(monto_pago)}, forma_pago={bool(forma_pago)}")
            return JsonResponse({
                'success': False,
                'message': 'Faltan campos obligatorios'
            })
        
        # Obtener PagoCliente
        pago_cliente = get_object_or_404(PagoCliente, id=pago_cliente_id)
        
        # Validar formato de fecha
        try:
            from datetime import datetime
            fecha_pago_obj = datetime.strptime(fecha_pago, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({
                'success': False,
                'message': f'Formato de fecha inválido: {fecha_pago}. Use formato YYYY-MM-DD'
            })
        
        # Procesar monto_pago - puede venir con formato (ej: "80.000")
        try:
            # Remover puntos y comas, luego convertir a float y luego a int
            monto_limpio = monto_pago.replace('.', '').replace(',', '')
            monto_pago_int = int(float(monto_limpio))
        except (ValueError, AttributeError):
            return JsonResponse({
                'success': False,
                'message': f'Formato de monto inválido: {monto_pago}. Use números sin formato (ej: 80000)'
            })
        
        # Validar que no se exceda el monto total
        if pago_cliente.monto_pagado + monto_pago_int > pago_cliente.monto_total:
            return JsonResponse({
                'success': False,
                'message': f'El monto excede el total pendiente (${pago_cliente.monto_pendiente:,})'
            })
        
        # Manejar archivo si se subió
        comprobante_pago = None
        if 'comprobante_pago' in request.FILES:
            comprobante_pago = request.FILES['comprobante_pago']
        
        # Crear DetallePago
        detalle_pago = DetallePago.objects.create(
            pago_cliente=pago_cliente,
            fecha_pago=fecha_pago_obj,  # Usar el objeto date validado
            monto_pago=monto_pago_int,
            forma_pago=forma_pago,
            numero_referencia=numero_referencia,
            comprobante_pago=comprobante_pago,
            notas=notas,
            registrado_por=request.user
        )
        
        # El modelo DetallePago ya actualiza automáticamente el monto_pagado en PagoCliente
        # Solo necesitamos refrescar el objeto para obtener los valores actualizados
        pago_cliente.refresh_from_db()
        
        return JsonResponse({
            'success': True,
            'message': 'Pago registrado exitosamente',
            'detalle_pago': {
                'id': detalle_pago.id,
                'fecha_pago': detalle_pago.fecha_pago.strftime('%d/%m/%Y'),
                'monto_pago': detalle_pago.monto_pago,
                'monto_pago_formateado': detalle_pago.monto_pago_formateado,
                'forma_pago': detalle_pago.get_forma_pago_display(),
                'numero_referencia': detalle_pago.numero_referencia or '',
                'notas': detalle_pago.notas or '',
                'fecha_registro': detalle_pago.fecha_registro.strftime('%d/%m/%Y %H:%M')
            },
            'pago_cliente_actualizado': {
                'monto_pagado': pago_cliente.monto_pagado,
                'monto_pendiente': pago_cliente.monto_pendiente,
                'estado_pago': pago_cliente.estado_pago,
                'porcentaje_pagado': pago_cliente.porcentaje_pagado
            }
        })
        
    except ValueError as e:
        return JsonResponse({
            'success': False,
            'message': f'Error en el formato de datos: {str(e)}'
        })
    except Exception as e:
        import traceback
        print(f"Error completo: {str(e)}")
        print(f"Traceback: {traceback.format_exc()}")
        return JsonResponse({
            'success': False,
            'message': f'Error al registrar pago: {str(e)}'
        })

@login_required
def obtener_historial_pagos_ajax(request, proyecto_id):
    """Vista AJAX para obtener historial de pagos de un proyecto"""
    try:
        proyecto = get_object_or_404(Proyecto, id=proyecto_id)
        pagos_cliente = PagoCliente.objects.filter(proyecto=proyecto).order_by('-fecha_creacion')
        
        historial = []
        for pago in pagos_cliente:
            detalles = []
            for detalle in pago.detalles_pago.all():
                detalles.append({
                    'fecha_pago': detalle.fecha_pago.strftime('%d/%m/%Y'),
                    'monto_pago': detalle.monto_pago_formateado,
                    'forma_pago': detalle.get_forma_pago_display(),
                    'numero_referencia': detalle.numero_referencia or '',
                    'notas': detalle.notas or '',
                    'fecha_registro': detalle.fecha_registro.strftime('%d/%m/%Y %H:%M')
                })
            
            historial.append({
                'id': pago.id,
                'cliente_nombre': pago.cliente_nombre,
                'monto_total': pago.monto_total_formateado,
                'monto_pagado': pago.monto_pagado_formateado,
                'monto_pendiente': pago.monto_pendiente_formateado,
                'estado_pago': pago.get_estado_pago_display(),
                'porcentaje_pagado': pago.porcentaje_pagado,
                'fecha_creacion': pago.fecha_creacion.strftime('%d/%m/%Y'),
                'fecha_vencimiento': pago.fecha_vencimiento.strftime('%d/%m/%Y') if pago.fecha_vencimiento else '',
                'descripcion': pago.descripcion or '',
                'observaciones': pago.observaciones or '',
                'detalles': detalles
            })
        
        return JsonResponse({
            'success': True,
            'proyecto': {
                'nombre': proyecto.nombre,
                'correo': proyecto.correo or '',
                'telefono': proyecto.telefono or ''
            },
            'historial': historial,
            'total_pagos': len(historial)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al obtener historial: {str(e)}'
        })

@login_required
def obtener_pagos_clientes_ajax(request):
    """Vista AJAX para obtener todos los pagos de clientes"""
    try:
        pagos_clientes = PagoCliente.objects.select_related('proyecto').prefetch_related('detalles_pago').order_by('-fecha_creacion')
        
        pagos_data = []
        for pago in pagos_clientes:
            pagos_data.append({
                'id': pago.id,
                'cliente_nombre': pago.cliente_nombre,
                'monto_total': pago.monto_total,
                'monto_pagado': pago.monto_pagado,
                'monto_pendiente': pago.monto_pendiente,
                'monto_total_formateado': pago.monto_total_formateado,
                'monto_pagado_formateado': pago.monto_pagado_formateado,
                'monto_pendiente_formateado': pago.monto_pendiente_formateado,
                'estado_pago': pago.get_estado_pago_display(),
                'porcentaje_pagado': pago.porcentaje_pagado,
                'fecha_creacion': pago.fecha_creacion.strftime('%d/%m/%Y'),
                'fecha_vencimiento': pago.fecha_vencimiento.strftime('%d/%m/%Y') if pago.fecha_vencimiento else '',
                'descripcion': pago.descripcion or '',
                'observaciones': pago.observaciones or '',
                'proyecto': {
                    'id': pago.proyecto.id,
                    'nombre': pago.proyecto.nombre,
                    'correo': pago.proyecto.correo or '',
                    'telefono': pago.proyecto.telefono or '',
                    'presupuesto': pago.proyecto.presupuesto or '',
                    'año': pago.proyecto.año
                } if pago.proyecto else None
            })
        
        return JsonResponse({
            'success': True,
            'pagos_clientes': pagos_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al obtener pagos de clientes: {str(e)}'
        })

@login_required
def obtener_info_pago_cliente_ajax(request, pago_cliente_id):
    """Vista AJAX para obtener información de un pago de cliente específico"""
    try:
        pago_cliente = get_object_or_404(PagoCliente, id=pago_cliente_id)
        
        return JsonResponse({
            'success': True,
            'pago_cliente': {
                'id': pago_cliente.id,
                'cliente_nombre': pago_cliente.cliente_nombre,
                'monto_total': pago_cliente.monto_total,
                'monto_pagado': pago_cliente.monto_pagado,
                'monto_pendiente': pago_cliente.monto_pendiente,
                'monto_total_formateado': pago_cliente.monto_total_formateado,
                'monto_pagado_formateado': pago_cliente.monto_pagado_formateado,
                'monto_pendiente_formateado': pago_cliente.monto_pendiente_formateado,
                'estado_pago': pago_cliente.get_estado_pago_display(),
                'porcentaje_pagado': pago_cliente.porcentaje_pagado,
                'fecha_creacion': pago_cliente.fecha_creacion.strftime('%d/%m/%Y'),
                'fecha_vencimiento': pago_cliente.fecha_vencimiento.strftime('%d/%m/%Y') if pago_cliente.fecha_vencimiento else '',
                'descripcion': pago_cliente.descripcion or '',
                'observaciones': pago_cliente.observaciones or ''
            }
        })
        
    except PagoCliente.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Pago de cliente no encontrado'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al obtener información del pago: {str(e)}'
        })

@login_required
def obtener_detalles_pago_cliente_ajax(request, pago_cliente_id):
    """Vista AJAX para obtener detalles de un pago de cliente"""
    try:
        pago_cliente = get_object_or_404(PagoCliente, id=pago_cliente_id)
        detalles_pago = pago_cliente.detalles_pago.all().order_by('-fecha_pago')
        
        detalles_data = []
        for detalle in detalles_pago:
            detalles_data.append({
                'id': detalle.id,
                'fecha_pago': detalle.fecha_pago.strftime('%d/%m/%Y'),
                'monto_pago': detalle.monto_pago,
                'monto_pago_formateado': detalle.monto_pago_formateado,
                'forma_pago': detalle.get_forma_pago_display(),
                'numero_referencia': detalle.numero_referencia or '',
                'notas': detalle.notas or '',
                'fecha_registro': detalle.fecha_registro.strftime('%d/%m/%Y %H:%M'),
                'comprobante_pago': detalle.comprobante_pago.url if detalle.comprobante_pago else ''
            })
        
        return JsonResponse({
            'success': True,
            'pago_cliente': {
                'id': pago_cliente.id,
                'cliente_nombre': pago_cliente.cliente_nombre,
                'monto_total': pago_cliente.monto_total,
                'monto_pagado': pago_cliente.monto_pagado,
                'monto_pendiente': pago_cliente.monto_pendiente,
                'monto_total_formateado': pago_cliente.monto_total_formateado,
                'monto_pagado_formateado': pago_cliente.monto_pagado_formateado,
                'monto_pendiente_formateado': pago_cliente.monto_pendiente_formateado,
                'estado_pago': pago_cliente.get_estado_pago_display(),
                'porcentaje_pagado': pago_cliente.porcentaje_pagado,
                'fecha_creacion': pago_cliente.fecha_creacion.strftime('%d/%m/%Y'),
                'fecha_vencimiento': pago_cliente.fecha_vencimiento.strftime('%d/%m/%Y') if pago_cliente.fecha_vencimiento else '',
                'descripcion': pago_cliente.descripcion or '',
                'observaciones': pago_cliente.observaciones or ''
            },
            'detalles_pago': detalles_data
        })
        
    except PagoCliente.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Pago de cliente no encontrado'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al obtener detalles del pago: {str(e)}'
        })

@login_required
def sistema_pagos_view(request):
    """Vista principal del sistema de pagos dinámico"""
    # Obtener años disponibles de los proyectos
    años_disponibles = Proyecto.objects.exclude(año__isnull=True).values_list('año', flat=True).distinct().order_by('-año')
    
    context = {
        'años_disponibles': años_disponibles
    }
    
    return render(request, 'pagos_cliente/sistema_pagos.html', context)

@login_required
def obtener_proyectos_con_pagos_ajax(request):
    """Vista AJAX para obtener todos los proyectos con información de pagos"""
    try:
        proyectos = Proyecto.objects.all().order_by('-id')
        
        proyectos_data = []
        for proyecto in proyectos:
            # Obtener información de pagos del proyecto
            pagos_cliente = PagoCliente.objects.filter(proyecto=proyecto).order_by('-fecha_creacion')
            
            pagos_info = None
            if pagos_cliente.exists():
                # Tomar el pago más reciente para mostrar en la tabla
                pago_principal = pagos_cliente.first()
                pago_principal.actualizar_estado()
                
                # Usar los valores del PagoCliente principal
                total_pagado = pago_principal.monto_pagado
                total_pendiente = pago_principal.monto_pendiente
                total_total = pago_principal.monto_total
                
                # Obtener último pago registrado
                ultimo_detalle = DetallePago.objects.filter(
                    pago_cliente__proyecto=proyecto
                ).order_by('-fecha_pago').first()
                
                pagos_info = {
                    'monto_total': total_total,
                    'monto_pagado': total_pagado,
                    'monto_pendiente': total_pendiente,
                    'monto_total_formateado': f"${total_total:,}",
                    'monto_pagado_formateado': f"${total_pagado:,}",
                    'monto_pendiente_formateado': f"${total_pendiente:,}",
                    'porcentaje_pagado': round((total_pagado / total_total) * 100) if total_total > 0 else 0,
                    'ultimo_pago': ultimo_detalle.fecha_pago.strftime('%d/%m/%Y') if ultimo_detalle else None,
                    'estado_pago': pago_principal.estado_pago
                }
            
            proyectos_data.append({
                'id': proyecto.id,
                'nombre': proyecto.nombre,
                'correo': proyecto.correo,
                'telefono': proyecto.telefono,
                'presupuesto': proyecto.presupuesto,
                'estado_proyecto': proyecto.estado_proyecto,
                'año': proyecto.año,
                'descripcion': proyecto.descripcion,
                'pagos_info': pagos_info
            })
        
        return JsonResponse({
            'success': True,
            'proyectos': proyectos_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al obtener proyectos: {str(e)}'
        })

@login_required
def obtener_info_proyecto_completa_ajax(request, proyecto_id):
    """Vista AJAX para obtener información completa de un proyecto"""
    try:
        proyecto = get_object_or_404(Proyecto, id=proyecto_id)
        pago_cliente = PagoCliente.objects.filter(proyecto=proyecto).first()
        
        proyecto_data = {
            'id': proyecto.id,
            'nombre': proyecto.nombre,
            'correo': proyecto.correo or '',
            'telefono': proyecto.telefono or '',
            'presupuesto': proyecto.presupuesto or '',
            'estado_proyecto': proyecto.estado_proyecto or '',
            'año': proyecto.año,
            'descripcion': proyecto.descripcion or ''
        }
        
        pago_cliente_data = None
        if pago_cliente:
            pago_cliente.actualizar_estado()
            pago_cliente_data = {
                'id': pago_cliente.id,
                'cliente_nombre': pago_cliente.cliente_nombre,
                'monto_total': pago_cliente.monto_total,
                'monto_pagado': pago_cliente.monto_pagado,
                'monto_pendiente': pago_cliente.monto_pendiente,
                'monto_total_formateado': pago_cliente.monto_total_formateado,
                'monto_pagado_formateado': pago_cliente.monto_pagado_formateado,
                'monto_pendiente_formateado': pago_cliente.monto_pendiente_formateado,
                'estado_pago': pago_cliente.get_estado_pago_display(),
                'porcentaje_pagado': pago_cliente.porcentaje_pagado
            }
        
        return JsonResponse({
            'success': True,
            'proyecto': proyecto_data,
            'pago_cliente': pago_cliente_data
        })
        
    except Proyecto.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Proyecto no encontrado'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al obtener información del proyecto: {str(e)}'
        })

@login_required
def obtener_detalles_proyecto_completos_ajax(request, proyecto_id):
    """Vista AJAX para obtener detalles completos de un proyecto con todos sus pagos"""
    try:
        proyecto = get_object_or_404(Proyecto, id=proyecto_id)
        pagos_cliente = PagoCliente.objects.filter(proyecto=proyecto).order_by('-fecha_creacion')
        
        proyecto_data = {
            'id': proyecto.id,
            'nombre': proyecto.nombre,
            'correo': proyecto.correo or '',
            'telefono': proyecto.telefono or '',
            'presupuesto': proyecto.presupuesto or '',
            'estado_proyecto': proyecto.estado_proyecto or '',
            'año': proyecto.año,
            'descripcion': proyecto.descripcion or ''
        }
        
        pagos_data = []
        for pago in pagos_cliente:
            pago.actualizar_estado()
            detalles = []
            
            for detalle in pago.detalles_pago.all():
                detalles.append({
                    'fecha_pago': detalle.fecha_pago.strftime('%d/%m/%Y'),
                    'monto_pago': detalle.monto_pago_formateado,
                    'forma_pago': detalle.get_forma_pago_display(),
                    'numero_referencia': detalle.numero_referencia or '',
                    'notas': detalle.notas or '',
                    'fecha_registro': detalle.fecha_registro.strftime('%d/%m/%Y %H:%M')
                })
            
            pagos_data.append({
                'id': pago.id,
                'cliente_nombre': pago.cliente_nombre,
                'monto_total': pago.monto_total,
                'monto_pagado': pago.monto_pagado,
                'monto_pendiente': pago.monto_pendiente,
                'monto_total_formateado': pago.monto_total_formateado,
                'monto_pagado_formateado': pago.monto_pagado_formateado,
                'monto_pendiente_formateado': pago.monto_pendiente_formateado,
                'estado_pago': pago.get_estado_pago_display(),
                'porcentaje_pagado': pago.porcentaje_pagado,
                'fecha_creacion': pago.fecha_creacion.strftime('%d/%m/%Y'),
                'fecha_vencimiento': pago.fecha_vencimiento.strftime('%d/%m/%Y') if pago.fecha_vencimiento else '',
                'descripcion': pago.descripcion or '',
                'observaciones': pago.observaciones or '',
                'detalles': detalles
            })
        
        return JsonResponse({
            'success': True,
            'proyecto': proyecto_data,
            'pagos_cliente': pagos_data
        })
        
    except Proyecto.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Proyecto no encontrado'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al obtener detalles del proyecto: {str(e)}'
        })

@login_required
@require_POST
def establecer_monto_total_ajax(request):
    proyecto_id = request.POST.get('proyecto_id')
    monto_total = request.POST.get('monto_total')
    descripcion = request.POST.get('descripcion', '')
    if not proyecto_id or not monto_total:
        return JsonResponse({'success': False, 'message': 'Datos incompletos.'})
    try:
        proyecto = Proyecto.objects.get(id=proyecto_id)
        if PagoCliente.objects.filter(proyecto=proyecto).exists():
            return JsonResponse({'success': False, 'message': 'Este proyecto ya tiene un monto total asignado.'})
        pago_cliente = PagoCliente.objects.create(
            proyecto=proyecto,
            cliente_nombre=proyecto.nombre,
            monto_total=monto_total,
            descripcion=descripcion or 'Monto total inicial',
            estado_pago='pendiente',
        )
        pago_cliente.actualizar_estado()
        return JsonResponse({'success': True, 'pago_cliente': {
            'id': pago_cliente.id,
            'monto_total': pago_cliente.monto_total,
        }})
    except Proyecto.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Proyecto no encontrado.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
def recalcular_montos_pagados(request):
    """Función para recalcular todos los montos pagados en caso de inconsistencias"""
    try:
        pagos_clientes = PagoCliente.objects.all()
        actualizados = 0
        
        for pago_cliente in pagos_clientes:
            # Calcular el monto pagado real sumando todos los detalles
            monto_pagado_real = pago_cliente.detalles_pago.aggregate(
                total=models.Sum('monto_pago')
            )['total'] or 0
            
            # Actualizar si hay diferencia
            if pago_cliente.monto_pagado != monto_pagado_real:
                pago_cliente.monto_pagado = monto_pagado_real
                pago_cliente.actualizar_estado()
                pago_cliente.save()
                actualizados += 1
        
        return JsonResponse({
            'success': True,
            'message': f'Se actualizaron {actualizados} registros de pagos de cliente'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al recalcular montos: {str(e)}'
        })

@login_required
def obtener_datos_proyecto_ajax(request, proyecto_id):
    try:
        proyecto = get_object_or_404(Proyecto, id=proyecto_id)
        data = {
            'id': proyecto.id,
            'presupuesto': proyecto.presupuesto,
            'año': proyecto.año,
            'nombre': proyecto.nombre,
            'telefono': proyecto.telefono,
            'correo': proyecto.correo,
            'direccion': proyecto.direccion,
            'rol_avaluo': proyecto.rol_avaluo,
            'proyecto': proyecto.proyecto,
            'estado_proyecto': proyecto.estado_proyecto,
            'proceso': proyecto.proceso,
            'kardex': proyecto.kardex,
            'detalle': proyecto.detalle,
            'descripcion': proyecto.descripcion,
            'notas_internas': proyecto.notas_internas,
            'presupuesto_link': proyecto.presupuesto_link,
            'enlace_documento': proyecto.enlace_documento,
            'link_1': proyecto.link_1,
            'nombre_enlace_1': proyecto.nombre_enlace_1,
            'levantamiento': proyecto.levantamiento.url if proyecto.levantamiento else None,
            'escritura': proyecto.escritura.url if proyecto.escritura else None,
            'dominio_vigente': proyecto.dominio_vigente.url if proyecto.dominio_vigente else None,
            'cedula_identidad': proyecto.cedula_identidad.url if proyecto.cedula_identidad else None,
            'poder': proyecto.poder.url if proyecto.poder else None,
            'boleta_agua': proyecto.boleta_agua.url if proyecto.boleta_agua else None,
            'boleta_luz': proyecto.boleta_luz.url if proyecto.boleta_luz else None,
            'rol_avaluo_detallado': proyecto.rol_avaluo_detallado.url if proyecto.rol_avaluo_detallado else None,
            'proyecto_anterior': proyecto.proyecto_anterior.url if proyecto.proyecto_anterior else None,
            'conjunto_habitacional': proyecto.conjunto_habitacional.url if proyecto.conjunto_habitacional else None,
            'informe_previo': proyecto.informe_previo.url if proyecto.informe_previo else None,
            'utilidad_publica': proyecto.utilidad_publica.url if proyecto.utilidad_publica else None,
            'certificado_numero': proyecto.certificado_numero.url if proyecto.certificado_numero else None,
            'factibilidad': proyecto.factibilidad.url if proyecto.factibilidad else None,
            'seim': proyecto.seim.url if proyecto.seim else None,
            'listado': proyecto.listado.url if proyecto.listado else None,
            'documento_presupuesto_1': proyecto.documento_presupuesto_1.url if proyecto.documento_presupuesto_1 else None,
            'documento_presupuesto_2': proyecto.documento_presupuesto_2.url if proyecto.documento_presupuesto_2 else None,
            'opcional_1': proyecto.opcional_1.url if proyecto.opcional_1 else None,
            'opcional_2': proyecto.opcional_2.url if proyecto.opcional_2 else None,
            'opcional_3': proyecto.opcional_3.url if proyecto.opcional_3 else None,
            'opcional_4': proyecto.opcional_4.url if proyecto.opcional_4 else None,
            'opcional_5': proyecto.opcional_5.url if proyecto.opcional_5 else None,
            'opcional_6': proyecto.opcional_6.url if proyecto.opcional_6 else None,
            'opcional_7': proyecto.opcional_7.url if proyecto.opcional_7 else None,
            'opcional_8': proyecto.opcional_8.url if proyecto.opcional_8 else None,
            'opcional_9': proyecto.opcional_9.url if proyecto.opcional_9 else None,
            'opcional_10': proyecto.opcional_10.url if proyecto.opcional_10 else None,
            'nombre_ingreso_1': proyecto.nombre_ingreso_1,
            'nombre_ingreso_2': proyecto.nombre_ingreso_2,
            'nombre_ingreso_3': proyecto.nombre_ingreso_3,
            'nombre_ingreso_4': proyecto.nombre_ingreso_4,
            'nombre_ingreso_5': proyecto.nombre_ingreso_5,
            'nombre_ingreso_6': proyecto.nombre_ingreso_6,
            'nombre_ingreso_7': proyecto.nombre_ingreso_7,
            'nombre_ingreso_8': proyecto.nombre_ingreso_8,
            'nombre_ingreso_9': proyecto.nombre_ingreso_9,
            'nombre_ingreso_10': proyecto.nombre_ingreso_10,
        }
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def proyecto_detail_json(request, proyecto_id):
    """Vista para obtener datos del proyecto en formato JSON para los modales"""
    try:
        proyecto = get_object_or_404(Proyecto, id=proyecto_id)
        data = {
            'id': proyecto.id,
            'presupuesto': proyecto.presupuesto,
            'año': proyecto.año,
            'nombre': proyecto.nombre,
            'telefono': proyecto.telefono,
            'correo': proyecto.correo,
            'direccion': proyecto.direccion,
            'rol_avaluo': proyecto.rol_avaluo,
            'proyecto': proyecto.proyecto,
            'estado_proyecto': proyecto.estado_proyecto,
            'proceso': proyecto.proceso,
            'kardex': proyecto.kardex,
            'detalle': proyecto.detalle,
            'descripcion': proyecto.descripcion,
            'notas_internas': proyecto.notas_internas,
            'presupuesto_link': proyecto.presupuesto_link,
            'enlace_documento': proyecto.enlace_documento,
            'link_1': proyecto.link_1,
            'nombre_enlace_1': proyecto.nombre_enlace_1,
            'levantamiento': proyecto.levantamiento.url if proyecto.levantamiento else None,
            'escritura': proyecto.escritura.url if proyecto.escritura else None,
            'dominio_vigente': proyecto.dominio_vigente.url if proyecto.dominio_vigente else None,
            'cedula_identidad': proyecto.cedula_identidad.url if proyecto.cedula_identidad else None,
            'poder': proyecto.poder.url if proyecto.poder else None,
            'boleta_agua': proyecto.boleta_agua.url if proyecto.boleta_agua else None,
            'boleta_luz': proyecto.boleta_luz.url if proyecto.boleta_luz else None,
            'rol_avaluo_detallado': proyecto.rol_avaluo_detallado.url if proyecto.rol_avaluo_detallado else None,
            'proyecto_anterior': proyecto.proyecto_anterior.url if proyecto.proyecto_anterior else None,
            'conjunto_habitacional': proyecto.conjunto_habitacional.url if proyecto.conjunto_habitacional else None,
            'informe_previo': proyecto.informe_previo.url if proyecto.informe_previo else None,
            'utilidad_publica': proyecto.utilidad_publica.url if proyecto.utilidad_publica else None,
            'certificado_numero': proyecto.certificado_numero.url if proyecto.certificado_numero else None,
            'factibilidad': proyecto.factibilidad.url if proyecto.factibilidad else None,
            'seim': proyecto.seim.url if proyecto.seim else None,
            'listado': proyecto.listado.url if proyecto.listado else None,
            'documento_presupuesto_1': proyecto.documento_presupuesto_1.url if proyecto.documento_presupuesto_1 else None,
            'documento_presupuesto_2': proyecto.documento_presupuesto_2.url if proyecto.documento_presupuesto_2 else None,
            'opcional_1': proyecto.opcional_1.url if proyecto.opcional_1 else None,
            'opcional_2': proyecto.opcional_2.url if proyecto.opcional_2 else None,
            'opcional_3': proyecto.opcional_3.url if proyecto.opcional_3 else None,
            'opcional_4': proyecto.opcional_4.url if proyecto.opcional_4 else None,
            'opcional_5': proyecto.opcional_5.url if proyecto.opcional_5 else None,
            'opcional_6': proyecto.opcional_6.url if proyecto.opcional_6 else None,
            'opcional_7': proyecto.opcional_7.url if proyecto.opcional_7 else None,
            'opcional_8': proyecto.opcional_8.url if proyecto.opcional_8 else None,
            'opcional_9': proyecto.opcional_9.url if proyecto.opcional_9 else None,
            'opcional_10': proyecto.opcional_10.url if proyecto.opcional_10 else None,
            'nombre_ingreso_1': proyecto.nombre_ingreso_1,
            'nombre_ingreso_2': proyecto.nombre_ingreso_2,
            'nombre_ingreso_3': proyecto.nombre_ingreso_3,
            'nombre_ingreso_4': proyecto.nombre_ingreso_4,
            'nombre_ingreso_5': proyecto.nombre_ingreso_5,
            'nombre_ingreso_6': proyecto.nombre_ingreso_6,
            'nombre_ingreso_7': proyecto.nombre_ingreso_7,
            'nombre_ingreso_8': proyecto.nombre_ingreso_8,
            'nombre_ingreso_9': proyecto.nombre_ingreso_9,
            'nombre_ingreso_10': proyecto.nombre_ingreso_10,
        }
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


